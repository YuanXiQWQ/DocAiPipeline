"""模板库路由：模板的增删改查、类型推荐、预览、下载。

模板库持久化到 SQLite templates 表，文件存放在 output/template_files/。
内置模板由 db.py 启动时自动注册，文件位于 models/ 目录。
"""

from __future__ import annotations

import json
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from loguru import logger
from pydantic import BaseModel

from app.config import settings
from app.db import get_conn

router = APIRouter(prefix="/api/template-lib", tags=["模板库"])

# 用户上传模板的存储目录
_TEMPLATE_FILES_DIR = Path(settings.output_dir) / "template_files"


def _ensure_dir() -> Path:
    _TEMPLATE_FILES_DIR.mkdir(parents=True, exist_ok=True)
    return _TEMPLATE_FILES_DIR


# ------------------------------------------------------------------
# Pydantic 模型
# ------------------------------------------------------------------


class TemplateRecord(BaseModel):
    """模板元数据。"""
    id: str
    name: str
    filename: str
    types: list[str]
    default_for: list[str]
    sheet_names: list[str]
    size_bytes: int
    builtin: bool
    imported_at: str
    last_used_at: str


class TemplatePreview(BaseModel):
    """模板预览：工作表列表 + 每个工作表的表头与前几行数据。"""
    sheet_names: list[str]
    sheets: dict[str, SheetPreview]


class SheetPreview(BaseModel):
    """单个工作表预览。"""
    headers: list[Any]
    rows: list[list[Any]]
    total_rows: int
    total_cols: int


# 重新定义以解决前向引用
TemplatePreview.model_rebuild()


class TypeRecommendation(BaseModel):
    """类型推荐结果。"""
    recommended: list[str]
    all_types: list[str]
    reason: str


# ------------------------------------------------------------------
# 已知工作表 → 文档类型的启发式映射
# ------------------------------------------------------------------

# 工作表名称关键词 → 推荐的文档类型
_SHEET_TYPE_HINTS: list[tuple[list[str], list[str]]] = [
    # 数据源表 → 原木/工厂类型
    (["数据源表"], ["log_measurement", "log_output", "soak_pool", "slicing", "packing"]),
    # 原始汇总 → 进口单据
    (["原始汇总"], ["customs"]),
    # 刨切/入池/上机 关键词
    (["刨切", "入池", "上机"], ["soak_pool", "slicing"]),
    # 打包
    (["打包", "表板"], ["packing"]),
    # 原木/检尺
    (["原木", "检尺", "码单"], ["log_measurement", "log_output"]),
]

# 表头关键词 → 推荐的文档类型
_HEADER_TYPE_HINTS: list[tuple[list[str], list[str]]] = [
    (["径级", "木种", "工序"], ["log_measurement", "log_output"]),
    (["入池", "池号"], ["soak_pool"]),
    (["刨切", "上机"], ["slicing"]),
    (["打包", "等级", "片数"], ["packing"]),
    (["供应商", "发票号", "外币"], ["customs"]),
]

# 所有已知文档类型
ALL_DOC_TYPES = ["customs", "log_measurement", "log_output", "soak_pool", "slicing", "packing"]


def _recommend_types(filepath: Path) -> tuple[list[str], str]:
    """根据 xlsx 结构启发式推荐适用的文档类型，返回 (推荐类型列表, 原因)。"""
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception:
        return [], "无法读取 xlsx 文件"

    sheet_names = wb.sheetnames
    reasons: list[str] = []
    matched: set[str] = set()

    # 按工作表名称匹配
    for keywords, types in _SHEET_TYPE_HINTS:
        for sn in sheet_names:
            if any(kw in sn for kw in keywords):
                matched.update(types)
                reasons.append(f"工作表「{sn}」匹配关键词 {keywords}")
                break

    # 按表头关键词匹配（检查每个工作表的前3行）
    for ws in wb.worksheets:
        header_texts: list[str] = []
        try:
            for r in range(1, 4):
                for c in range(1, min(ws.max_column or 1, 50) + 1):
                    val = ws.cell(r, c).value
                    if isinstance(val, str):
                        header_texts.append(val)
        except Exception:
            continue

        combined = " ".join(header_texts)
        for keywords, types in _HEADER_TYPE_HINTS:
            if any(kw in combined for kw in keywords):
                matched.update(types)
                if not any(kw in " ".join(reasons) for kw in keywords):
                    reasons.append(f"工作表「{ws.title}」表头含关键词 {keywords}")

    wb.close()

    reason = "；".join(reasons) if reasons else "未检测到已知特征，建议手动选择"
    return sorted(matched), reason


def _read_sheet_names(filepath: Path) -> list[str]:
    """读取 xlsx 的工作表名称列表。"""
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def _row_to_record(row: Any) -> TemplateRecord:
    """将 SQLite Row 转换为 TemplateRecord。"""
    return TemplateRecord(
        id=row["id"],
        name=row["name"],
        filename=row["filename"],
        types=json.loads(row["types"]),
        default_for=json.loads(row["default_for"]),
        sheet_names=json.loads(row["sheet_names"]),
        size_bytes=row["size_bytes"],
        builtin=bool(row["builtin"]),
        imported_at=row["imported_at"],
        last_used_at=row["last_used_at"],
    )


# ------------------------------------------------------------------
# 列表 & 详情
# ------------------------------------------------------------------


@router.get("", response_model=list[TemplateRecord])
async def list_templates(sort_by: str = "last_used_at", category: str | None = None):
    """列出所有模板。

    - sort_by: last_used_at（默认）| imported_at | name
    - category: 按文档类型过滤（可选）
    """
    conn = get_conn()

    order_col = "last_used_at"
    if sort_by == "imported_at":
        order_col = "imported_at"
    elif sort_by == "name":
        order_col = "name"

    if category:
        # JSON 数组中包含某类型的过滤（SQLite JSON 函数）
        rows = conn.execute(
            f"""SELECT * FROM templates
                WHERE types LIKE ?
                ORDER BY {order_col} DESC""",
            (f'%"{category}"%',),
        ).fetchall()
    else:
        rows = conn.execute(
            f"SELECT * FROM templates ORDER BY {order_col} DESC"
        ).fetchall()

    return [_row_to_record(r) for r in rows]


@router.get("/{template_id}", response_model=TemplateRecord)
async def get_template(template_id: str):
    """获取单个模板详情。"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")
    return _row_to_record(row)


# ------------------------------------------------------------------
# 预览
# ------------------------------------------------------------------


@router.get("/{template_id}/preview")
async def preview_template(template_id: str, max_rows: int = 5):
    """预览模板：每个工作表的表头 + 前几行数据（纯值，不含公式）。"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")

    filepath = Path(row["file_path"])
    if not filepath.exists():
        raise HTTPException(404, "模板文件不存在")

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(500, f"读取模板失败: {e}")

    sheets: dict[str, Any] = {}
    for ws in wb.worksheets:
        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        headers: list[Any] = []
        rows_data: list[list[Any]] = []

        # 读取表头（第1行或第2行）
        for r in range(1, min(3, total_rows + 1)):
            row_vals = []
            for c in range(1, min(total_cols + 1, 50)):
                v = ws.cell(r, c).value
                row_vals.append(v)
            if r == 1:
                headers = row_vals
            else:
                rows_data.append(row_vals)

        # 读取数据行
        for r in range(3, min(3 + max_rows, total_rows + 1)):
            row_vals = []
            for c in range(1, min(total_cols + 1, 50)):
                v = ws.cell(r, c).value
                row_vals.append(v)
            rows_data.append(row_vals)

        sheets[ws.title] = {
            "headers": headers,
            "rows": rows_data,
            "total_rows": total_rows,
            "total_cols": total_cols,
        }

    wb.close()
    return {"sheet_names": list(sheets.keys()), "sheets": sheets}


# ------------------------------------------------------------------
# 导入（上传）
# ------------------------------------------------------------------


@router.post("/recommend-types", response_model=TypeRecommendation)
async def recommend_types(file: UploadFile = File(...)):
    """上传 xlsx 文件，返回推荐的适用文档类型（不保存文件）。"""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "请上传 .xlsx 文件")

    # 保存到临时路径分析
    tmp_dir = _ensure_dir()
    tmp_path = tmp_dir / f"_tmp_{uuid.uuid4().hex[:6]}.xlsx"
    try:
        with open(tmp_path, "wb") as buf:
            buf.write(await file.read())
        recommended, reason = _recommend_types(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)

    return TypeRecommendation(
        recommended=recommended,
        all_types=ALL_DOC_TYPES,
        reason=reason,
    )


@router.post("", response_model=TemplateRecord)
async def import_template(
    file: UploadFile = File(...),
    name: str = "",
    types_json: str = "[]",
):
    """导入模板文件到模板库。

    - file: xlsx 文件
    - name: 显示名称（可选，默认取文件名）
    - types_json: 适用类型 JSON 数组字符串
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "请上传 .xlsx 文件")

    try:
        types: list[str] = json.loads(types_json)
    except json.JSONDecodeError:
        types = []

    # 保存文件
    store_dir = _ensure_dir()
    tpl_id = uuid.uuid4().hex[:12]
    stored_name = f"{tpl_id}_{file.filename}"
    stored_path = store_dir / stored_name

    content = await file.read()
    with open(stored_path, "wb") as buf:
        buf.write(content)

    # 如果没有指定类型，自动推荐
    if not types:
        types, _ = _recommend_types(stored_path)

    # 读取工作表名称
    sheet_names = _read_sheet_names(stored_path)

    display_name = name or Path(file.filename).stem
    now = datetime.now(timezone.utc).isoformat()

    conn = get_conn()
    conn.execute(
        """INSERT INTO templates
           (id, name, filename, file_path, types, default_for,
            sheet_names, size_bytes, builtin, imported_at, last_used_at)
           VALUES (?,?,?,?,?,?,?,?,0,?,?)""",
        (
            tpl_id,
            display_name,
            file.filename,
            str(stored_path),
            json.dumps(types, ensure_ascii=False),
            "[]",
            json.dumps(sheet_names, ensure_ascii=False),
            len(content),
            now,
            "",
        ),
    )
    conn.commit()

    logger.info(f"模板已导入: {display_name} ({file.filename}), 类型: {types}")

    row = conn.execute("SELECT * FROM templates WHERE id = ?", (tpl_id,)).fetchone()
    return _row_to_record(row)


# ------------------------------------------------------------------
# 更新
# ------------------------------------------------------------------


@router.put("/{template_id}", response_model=TemplateRecord)
async def update_template(template_id: str, body: dict):
    """更新模板元数据（名称、适用类型）。"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")

    updates: list[str] = []
    params: list[Any] = []

    if "name" in body:
        updates.append("name = ?")
        params.append(body["name"])
    if "types" in body:
        updates.append("types = ?")
        params.append(json.dumps(body["types"], ensure_ascii=False))

    if not updates:
        raise HTTPException(400, "无可更新字段")

    params.append(template_id)
    conn.execute(
        f"UPDATE templates SET {', '.join(updates)} WHERE id = ?",
        params,
    )
    conn.commit()

    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    return _row_to_record(row)


@router.post("/{template_id}/set-default")
async def set_default(template_id: str, body: dict):
    """设置/取消模板为某类型的默认模板。

    body: { "doc_type": "log_measurement", "is_default": true }
    每种类型只能有一个默认模板。
    """
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")

    doc_type = body.get("doc_type", "")
    is_default = body.get("is_default", True)

    if not doc_type:
        raise HTTPException(400, "缺少 doc_type")

    if is_default:
        # 先移除该类型在所有模板中的默认标记
        all_rows = conn.execute("SELECT id, default_for FROM templates").fetchall()
        for r in all_rows:
            df: list[str] = json.loads(r["default_for"])
            if doc_type in df:
                df.remove(doc_type)
                conn.execute(
                    "UPDATE templates SET default_for = ? WHERE id = ?",
                    (json.dumps(df, ensure_ascii=False), r["id"]),
                )

        # 为当前模板添加默认标记
        current_df: list[str] = json.loads(row["default_for"])
        if doc_type not in current_df:
            current_df.append(doc_type)
        conn.execute(
            "UPDATE templates SET default_for = ? WHERE id = ?",
            (json.dumps(current_df, ensure_ascii=False), template_id),
        )
    else:
        # 移除当前模板的该类型默认标记
        current_df = json.loads(row["default_for"])
        if doc_type in current_df:
            current_df.remove(doc_type)
            conn.execute(
                "UPDATE templates SET default_for = ? WHERE id = ?",
                (json.dumps(current_df, ensure_ascii=False), template_id),
            )

    conn.commit()

    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    return _row_to_record(row)


@router.post("/{template_id}/touch")
async def touch_template(template_id: str):
    """更新模板的最近使用时间。"""
    conn = get_conn()
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        "UPDATE templates SET last_used_at = ? WHERE id = ?",
        (now, template_id),
    )
    conn.commit()
    return {"ok": True}


# ------------------------------------------------------------------
# 下载
# ------------------------------------------------------------------


@router.get("/{template_id}/download")
async def download_template(template_id: str):
    """下载模板文件。"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")

    filepath = Path(row["file_path"])
    if not filepath.exists():
        raise HTTPException(404, "模板文件不存在")

    return FileResponse(
        path=str(filepath),
        filename=row["filename"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ------------------------------------------------------------------
# 删除
# ------------------------------------------------------------------


@router.delete("/{template_id}")
async def delete_template(template_id: str):
    """删除模板（包括文件）。"""
    conn = get_conn()
    row = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(404, "模板不存在")

    # 删除文件（内置模板不删除源文件）
    if not row["builtin"]:
        filepath = Path(row["file_path"])
        if filepath.exists():
            filepath.unlink(missing_ok=True)

    conn.execute("DELETE FROM templates WHERE id = ?", (template_id,))
    conn.commit()

    logger.info(f"模板已删除: {row['name']} ({row['filename']})")
    return {"ok": True, "message": f"已删除模板: {row['name']}"}


# ------------------------------------------------------------------
# 按文档类型查找匹配模板（供填充流程使用）
# ------------------------------------------------------------------


@router.get("/match/{doc_type}", response_model=list[TemplateRecord])
async def find_matching_templates(doc_type: str):
    """查找适用于指定文档类型的所有模板，默认模板排在最前。"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT * FROM templates
           WHERE types LIKE ?
           ORDER BY
               CASE WHEN default_for LIKE ? THEN 0 ELSE 1 END,
               last_used_at DESC""",
        (f'%"{doc_type}"%', f'%"{doc_type}"%'),
    ).fetchall()
    return [_row_to_record(r) for r in rows]
