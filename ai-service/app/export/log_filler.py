"""检尺单数据填充器：将识别结果写入「原木出入库-3-3.xlsx」数据源表。

由于目标 Excel 含透视表，openpyxl 会因 CalculatedItem.formula=None 而崩溃。
解决方案：monkey-patch openpyxl 的 CalculatedItem 使 formula 字段可选。
"""

from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from app.schemas import LogEntry, LogMeasurementResult, LogSheetMeta

# ------------------------------------------------------------------
# Monkey-patch: 修复 openpyxl 透视表 CalculatedItem.formula 必须为 str 的 bug
# ------------------------------------------------------------------

def _patch_openpyxl_pivot_cache() -> None:
    """修复 openpyxl 透视表加载问题。

    1. 让 CalculatedItem.formula 接受 None，避免 TypeError。
    2. 跳过 pivot_caches 解析，避免读取巨大的缓存文件导致卡住。
    """
    try:
        from openpyxl.pivot.cache import CalculatedItem
        from openpyxl.descriptors.base import String
        # 将 formula 改为 allow_none=True
        CalculatedItem.formula = String(allow_none=True)  # type: ignore[assignment]
    except (ImportError, AttributeError):
        pass  # 如果 openpyxl 版本不同，跳过

    try:
        from openpyxl.reader.workbook import WorkbookParser

        # 跳过 pivotCacheRecords 解析（可达数十 MB），返回空字典
        # noinspection PyPropertyAccess
        WorkbookParser.pivot_caches = property(lambda _: {})  # type: ignore[assignment]

        # read_worksheets 中 pivot_caches[cacheId] 会 KeyError，需要跳过
        import openpyxl.reader.excel as _xl_mod
        _orig_read_ws = _xl_mod.ExcelReader.read_worksheets

        def _safe_read_worksheets(self: object) -> None:  # type: ignore[no-untyped-def]
            """包装 read_worksheets：让 pivot 关联失败时跳过而非中断。"""
            from openpyxl.xml.functions import fromstring
            from openpyxl.packaging.relationship import (
                get_dependents, get_rels_path, RelationshipList,
            )
            from openpyxl.worksheet._reader import WorksheetReader
            from openpyxl.worksheet._read_only import ReadOnlyWorksheet
            from openpyxl.comments.comment_sheet import CommentSheet
            from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
            from openpyxl.reader.drawings import find_images
            from openpyxl.worksheet.table import Table

            comments_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments"  # noqa: S105

            for sheet, rel in self.parser.find_sheets():  # type: ignore[attr-defined]
                if rel.target not in self.valid_files:  # type: ignore[attr-defined]
                    continue
                if "chartsheet" in rel.Type:
                    self.read_chartsheet(sheet, rel)  # type: ignore[attr-defined]
                    continue

                rels_path = get_rels_path(rel.target)
                rels = RelationshipList()
                if rels_path in self.valid_files:  # type: ignore[attr-defined]
                    rels = get_dependents(self.archive, rels_path)  # type: ignore[attr-defined]

                if self.read_only:  # type: ignore[attr-defined]
                    ws = ReadOnlyWorksheet(self.wb, sheet.name, rel.target, self.shared_strings)  # type: ignore[attr-defined]
                    ws.sheet_state = sheet.state
                    # noinspection PyProtectedMember
                    self.wb._sheets.append(ws)  # type: ignore[attr-defined]  # noqa: SLF001
                    continue

                fh = self.archive.open(rel.target)  # type: ignore[attr-defined]
                ws = self.wb.create_sheet(sheet.name)  # type: ignore[attr-defined]
                ws._rels = rels
                ws_parser = WorksheetReader(ws, fh, self.shared_strings, self.data_only, self.rich_text)  # type: ignore[attr-defined]
                ws_parser.bind_all()
                fh.close()

                for r in rels.find(comments_ns):
                    src = self.archive.read(r.target)  # type: ignore[attr-defined]
                    comment_sheet = CommentSheet.from_tree(fromstring(src))  # type: ignore[arg-type]
                    for ref, comment in comment_sheet.comments:
                        try:
                            ws[ref].comment = comment
                        except AttributeError:
                            pass

                if self.wb.vba_archive and ws.legacy_drawing:  # type: ignore[attr-defined]
                    ws.legacy_drawing = rels.get(ws.legacy_drawing).target
                else:
                    ws.legacy_drawing = None

                for t in ws_parser.tables:
                    src = self.archive.read(t)  # type: ignore[attr-defined]
                    xml = fromstring(src)
                    table = Table.from_tree(xml)  # type: ignore[arg-type]
                    ws.add_table(table)

                # noinspection PyProtectedMember
                drawings = rels.find(SpreadsheetDrawing._rel_type)  # type: ignore[attr-defined]  # noqa: SLF001
                for r in drawings:
                    charts, images = find_images(self.archive, r.target)  # type: ignore[attr-defined]
                    for c in charts:
                        ws.add_chart(c, c.anchor)
                    for im in images:
                        ws.add_image(im, im.anchor)

                # ★ 跳过 pivot 关联（避免 KeyError / 巨大缓存解析）

                ws.sheet_state = sheet.state

        _xl_mod.ExcelReader.read_worksheets = _safe_read_worksheets  # type: ignore[assignment]
    except (ImportError, AttributeError):
        pass

_patch_openpyxl_pivot_cache()


# ------------------------------------------------------------------
# 数据源表列映射（0-indexed → 1-indexed for openpyxl）
# ------------------------------------------------------------------

# 数据源表列定义（col index 1-based）
COL_YEAR = 1           # A: 年
COL_MONTH = 2          # B: 月份
COL_DATE = 3           # C: 日期
COL_WORKSHOP = 4       # D: 车间
COL_MATERIAL = 5       # E: 物料性质
COL_PROCESS = 6        # F: 工序
COL_SPECIES = 7        # G: 木种
COL_GOODS_NAME = 8     # H: 货物名称
COL_BATCH_ID = 9       # I: 序号/批次号
COL_CUSTOMER = 10      # J: 送货客户
COL_VEHICLE = 11       # K: 车牌号
COL_OWNER = 12         # L: 货物所有人
COL_LOG_ID = 13        # M: 包号/根号
COL_GRADE = 14         # N: 等级
COL_CRAFT = 15         # O: 工艺
COL_SAW_NUM = 16       # P: 锯号/池号
COL_INPUT_M3 = 17      # Q: 投入数量 m3
COL_LENGTH = 18        # R: 长度
COL_DIAMETER = 19      # S: 宽度/径级
COL_THICKNESS = 20     # T: 厚度
COL_CALC_LENGTH = 21   # U: 计尺长
COL_CALC_WIDTH = 22    # V: 计尺宽
COL_CALC_THICK = 23    # W: 计尺厚
COL_COUNT = 24         # X: 片数/根数
COL_QTY_M2 = 25        # Y: 数量 m2
COL_QTY_M3 = 26        # Z: 数量 m3
COL_PACK_NUM = 27      # AA: 包数
COL_PACK_QTY = 28      # AB: 每包数量
COL_COL1 = 29          # AC: 列1
COL_WEIGHT = 30        # AD: 重量

# 表头行（数据从第3行开始，第1行标签，第2行列名）
HEADER_ROWS = 2
DATA_START_ROW = 3


class LogFiller:
    """将检尺单识别结果填入「原木出入库」数据源表。"""

    SHEET_NAME = "数据源表"

    def __init__(
        self,
        template_path: Path,
        *,
        workshop: str = "大锯车间",
        material: str = "原料",
        process: str = "来料",
        species: str = "欧橡",
        goods_name: str = "原木",
        owner: str = "新公司",
    ):
        self.template_path = template_path
        self.workshop = workshop
        self.material = material
        self.process = process
        self.species = species
        self.goods_name = goods_name
        self.owner = owner

        if not self.template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {self.template_path}")

    def fill(
        self,
        results: list[LogMeasurementResult],
        output_path: Path,
        *,
        grade: str = "",
        customer: str = "",
    ) -> Path:
        """将一组检尺单识别结果写入 Excel 数据源表。

        参数:
            results: LogExtractor 的识别结果列表。
            output_path: 输出文件路径。
            grade: 等级（如 "厚皮"、"F1" 等），默认空。
            customer: 送货客户，默认空。

        返回:
            保存的文件路径。
        """
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb[self.SHEET_NAME]

        # 找到第一个空行
        start_row = self._find_first_empty_row(ws)
        logger.info(
            f"数据源表已有数据到第 {start_row - 1} 行，从第 {start_row} 行开始追加。"
        )

        current_row = start_row
        total_entries = 0

        for result in results:
            # 只处理有逐行数据的页面（跳过确认表和仅编号表）
            if not result.entries:
                continue

            # 跳过 length=0 的行（仅编号列表）
            valid_entries = [
                e for e in result.entries
                if e.length_m > 0 and e.diameter_cm > 0
            ]
            if not valid_entries:
                continue

            # 解析日期
            date_obj = self._parse_date(result.meta.date)

            for entry in valid_entries:
                volume = self._calc_volume(entry)
                self._write_row(
                    ws,
                    row=current_row,
                    date_obj=date_obj,
                    meta=result.meta,
                    entry=entry,
                    volume=volume,
                    grade=grade,
                    customer=customer or result.meta.supplier,
                )
                current_row += 1
                total_entries += 1

        logger.info(f"共写入 {total_entries} 行到 '{self.SHEET_NAME}'")

        # 保存
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info(f"已保存: {output_path}")

        return output_path

    # ------------------------------------------------------------------
    # 私有方法
    # ------------------------------------------------------------------

    @staticmethod
    def _find_first_empty_row(ws: Any) -> int:
        """找到数据源表中第一个空行。"""
        for row in range(DATA_START_ROW, ws.max_row + 2):
            # 检查 A 列（年）是否有数据
            cell_val = ws.cell(row=row, column=COL_YEAR).value
            if cell_val is None or cell_val == "":
                return row
            # 也检查年份是否为占位值（如 1900）
            try:
                if int(cell_val) < 2000:
                    return row
            except (ValueError, TypeError):
                pass
        return ws.max_row + 1

    @staticmethod
    def _parse_date(date_str: str) -> datetime | None:
        """解析各种格式的日期字符串。"""
        if not date_str:
            return None

        # 格式: "26年2月6日" → 2026-02-06
        import re
        m = re.match(r"(\d{2,4})\u5e74(\d{1,2})\u6708(\d{1,2})\u65e5?", date_str)
        if m:
            year = int(m.group(1))
            if year < 100:
                year += 2000
            return datetime(year, int(m.group(2)), int(m.group(3)))

        # 格式: "YYYY-MM-DD"
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            pass

        # 格式: "DD.MM.YYYY"
        try:
            return datetime.strptime(date_str, "%d.%m.%Y")
        except ValueError:
            pass

        logger.warning(f"无法解析日期: {date_str}")
        return None

    @staticmethod
    def _calc_volume(entry: LogEntry) -> float:
        """计算单根原木体积 (m³)。

        如果 VLM 已返回体积（打印码单），直接使用。
        否则用简化 JAS 公式：V = π/4 × D² × L
        """
        if entry.volume_m3 is not None:
            return entry.volume_m3
        d_m = entry.diameter_cm / 100.0
        return round(math.pi / 4 * d_m ** 2 * entry.length_m, 2)

    def _write_row(
        self,
        ws: Any,
        *,
        row: int,
        date_obj: datetime | None,
        meta: LogSheetMeta,
        entry: LogEntry,
        volume: float,
        grade: str,
        customer: str,
    ) -> None:
        """向数据源表写入一行数据。"""
        if date_obj:
            ws.cell(row=row, column=COL_YEAR, value=date_obj.year)
            ws.cell(row=row, column=COL_MONTH, value=date_obj.month)
            ws.cell(row=row, column=COL_DATE, value=date_obj)

        ws.cell(row=row, column=COL_WORKSHOP, value=self.workshop)
        ws.cell(row=row, column=COL_MATERIAL, value=self.material)
        ws.cell(row=row, column=COL_PROCESS, value=self.process)
        ws.cell(row=row, column=COL_SPECIES, value=self.species)
        ws.cell(row=row, column=COL_GOODS_NAME, value=self.goods_name)
        ws.cell(row=row, column=COL_BATCH_ID, value=meta.batch_id)
        ws.cell(row=row, column=COL_CUSTOMER, value=customer)
        ws.cell(row=row, column=COL_VEHICLE, value=meta.vehicle_plate)
        ws.cell(row=row, column=COL_OWNER, value=self.owner)
        ws.cell(row=row, column=COL_LOG_ID, value=entry.log_id)
        ws.cell(row=row, column=COL_GRADE, value=grade)

        # 长度和径级
        ws.cell(row=row, column=COL_LENGTH, value=entry.length_m)
        ws.cell(row=row, column=COL_DIAMETER, value=entry.diameter_cm)

        # 计尺尺寸 = 原始尺寸（原木无厚度）
        ws.cell(row=row, column=COL_CALC_LENGTH, value=entry.length_m)
        ws.cell(row=row, column=COL_CALC_WIDTH, value=entry.diameter_cm)
        ws.cell(row=row, column=COL_CALC_THICK, value=0)

        # 数量
        ws.cell(row=row, column=COL_QTY_M2, value=0)
        ws.cell(row=row, column=COL_QTY_M3, value=volume)

        # 重量 = 体积（用于透视表汇总）
        ws.cell(row=row, column=COL_WEIGHT, value=volume)
