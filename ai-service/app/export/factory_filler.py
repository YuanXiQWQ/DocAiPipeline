"""工厂内部单据 Excel 填充器。

所有文档类型均写入统一的「数据统计」模板的「数据源表」工作表。

数据源表列布局（R2 表头）：
  C1=年(公式)  C2=月(公式)  C3=日期  C4=车间  C5=物料性质
  C6=工序  C7=木种  C8=货物名称  C9=序号/批次号  C10=送货客户
  C11=车牌号  C12=货物所有人  C13=包号/根号  C14=等级  C15=工艺
  C16=锯号/池号  C17=投入数量m³  C18=长度  C19=宽度/径级  C20=厚度
  C21=计尺长(公式)  C22=计尺宽(公式)  C23=计尺厚(公式)  C24=片数/根数
  C25=m²(公式)  C26=m³(公式)  ...  C35=检索规格(公式)

公式列（C1/C2/C21-C23/C25-C26/C35）由模板预填，不可覆写。

复用 log_filler.py 中的 monkey-patch 来绕过透视表问题。
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from app.schemas import (
    LogOutputResult,
    PackingResult,
    SlicingResult,
    SoakPoolResult,
)

# 确保 monkey-patch 已执行
from app.export.log_filler import _patch_openpyxl_pivot_cache

_patch_openpyxl_pivot_cache()


# ------------------------------------------------------------------
# 通用工具
# ------------------------------------------------------------------

def _parse_date(date_str: str) -> datetime | None:
    """解析各种格式的日期字符串。"""
    if not date_str:
        return None
    # "26年2月6日"
    m = re.match(r"(\d{2,4})\u5e74(\d{1,2})\u6708(\d{1,2})\u65e5?", date_str)
    if m:
        y = int(m.group(1))
        if y < 100:
            y += 2000
        return datetime(y, int(m.group(2)), int(m.group(3)))
    # "YYYY-MM-DD"
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        pass
    # "DD-MM-YYYY"
    try:
        return datetime.strptime(date_str, "%d-%m-%Y")
    except ValueError:
        pass
    # "26-2-2024" (可能是 YY-M-YYYY 误格式)
    m2 = re.match(r"(\d{2})-(\d{1,2})-(\d{4})", date_str)
    if m2:
        return datetime(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    logger.warning(f"无法解析日期: {date_str}")
    return None


def _check_sheet(wb: Any, name: str, template_path: Path) -> Any:
    """校验目标工作表是否存在，不存在时给出清晰提示。"""
    if name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"模板文件 '{template_path.name}' 中不存在工作表 '{name}'。"
            f"可用工作表: [{available}]"
        )
    return wb[name]


# 数据源表中的公式列，不可覆写
FORMULA_COLS = frozenset({1, 2, 21, 22, 23, 25, 26, 35})


def _find_first_empty_row(ws: Any, date_col: int = 3, start_row: int = 3) -> int:
    """找到数据源表中第一个空行（基于日期列 C3）。"""
    for row in range(start_row, ws.max_row + 2):
        val = ws.cell(row=row, column=date_col).value
        if val is None or val == "":
            return row
    return ws.max_row + 1


# ------------------------------------------------------------------
# 1. 出库表 → 数据源表（工序=领用出库）
# ------------------------------------------------------------------

class LogOutputFiller:
    """原木领用出库表 → 数据源表（工序=领用出库）。"""

    SHEET_NAME = "数据源表"

    def __init__(self, template_path: Path):
        self.template_path = template_path

    def fill(
            self, results: list[LogOutputResult], output_path: Path,
            *, species: str = "欧橡", owner: str = "新公司",
    ) -> Path:
        wb = openpyxl.load_workbook(self.template_path)
        ws = _check_sheet(wb, self.SHEET_NAME, self.template_path)
        start = _find_first_empty_row(ws)
        current = start
        count = 0

        for r in results:
            date_obj = _parse_date(r.meta.date)
            for e in r.entries:
                if e.diameter_cm <= 0:
                    continue
                if date_obj:
                    ws.cell(current, 3, date_obj)         # 日期
                ws.cell(current, 4, "大锯车间")            # 车间
                ws.cell(current, 5, "原料")                # 物料性质
                ws.cell(current, 6, "领用出库")            # 工序
                ws.cell(current, 7, species)               # 木种
                ws.cell(current, 8, "原木")                # 货物名称
                ws.cell(current, 9, r.meta.batch_id)       # 序号/批次号
                ws.cell(current, 12, owner)                # 货物所有人
                ws.cell(current, 13, e.log_id)             # 包号/根号
                ws.cell(current, 19, e.diameter_cm)        # 宽度/径级
                ws.cell(current, 24, 1)                    # 根数
                current += 1
                count += 1

        logger.info(f"出库表写入 {count} 行到 '{self.SHEET_NAME}'")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path


# ------------------------------------------------------------------
# 2. 入池表 → 数据源表（工序=入池）
# ------------------------------------------------------------------

class SoakPoolFiller:
    """刨切木方入池表 → 数据源表（工序=入池）。"""

    SHEET_NAME = "数据源表"

    def __init__(self, template_path: Path):
        self.template_path = template_path

    def fill(
            self, results: list[SoakPoolResult], output_path: Path,
            *, owner: str = "新公司",
    ) -> Path:
        wb = openpyxl.load_workbook(self.template_path)
        ws = _check_sheet(wb, self.SHEET_NAME, self.template_path)
        start = _find_first_empty_row(ws)
        current = start
        count = 0

        for r in results:
            date_obj = _parse_date(r.meta.date)
            for e in r.entries:
                if date_obj:
                    ws.cell(current, 3, date_obj)                      # 日期
                ws.cell(current, 5, "半成品")                           # 物料性质
                ws.cell(current, 6, "入池")                             # 工序
                ws.cell(current, 7, "欧橡")                             # 木种
                ws.cell(current, 8, "大方")                             # 货物名称
                ws.cell(current, 9, r.meta.batch_id)                   # 序号/批次号
                ws.cell(current, 12, r.meta.owner or owner)            # 货物所有人
                ws.cell(current, 15, r.meta.craft or "刨切")            # 工艺
                ws.cell(current, 16, r.meta.pool_number)               # 池号
                ws.cell(current, 18, e.length_mm)                      # 长度
                ws.cell(current, 19, e.width_mm)                       # 宽度
                ws.cell(current, 20, e.thickness_mm)                   # 厚度
                ws.cell(current, 24, 1)                                # 根数
                current += 1
                count += 1

        logger.info(f"入池表写入 {count} 行到 '{self.SHEET_NAME}'")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path


# ------------------------------------------------------------------
# 3. 上机表 → 数据源表（工序=生产/刨切）
# ------------------------------------------------------------------

class SlicingFiller:
    """刨切上机表 → 数据源表（工序=生产）。"""

    SHEET_NAME = "数据源表"

    def __init__(self, template_path: Path):
        self.template_path = template_path

    def fill(
            self, results: list[SlicingResult], output_path: Path,
            *, owner: str = "新公司",
    ) -> Path:
        wb = openpyxl.load_workbook(self.template_path)
        ws = _check_sheet(wb, self.SHEET_NAME, self.template_path)
        start = _find_first_empty_row(ws)
        current = start
        count = 0

        for r in results:
            date_obj = _parse_date(r.meta.date)
            for e in r.entries:
                if date_obj:
                    ws.cell(current, 3, date_obj)                      # 日期
                ws.cell(current, 5, "半成品")                           # 物料性质
                ws.cell(current, 6, "生产")                             # 工序
                ws.cell(current, 7, "欧橡")                             # 木种
                ws.cell(current, 8, "刨切表板")                         # 货物名称
                ws.cell(current, 9, r.meta.batch_id)                   # 序号/批次号
                ws.cell(current, 12, r.meta.owner or owner)            # 货物所有人
                ws.cell(current, 15, "刨切")                            # 工艺
                ws.cell(current, 19, e.width_mm)                       # 宽度
                ws.cell(current, 20, e.thickness_mm)                   # 厚度（大方厚）
                ws.cell(current, 24, 1)                                # 根数
                current += 1
                count += 1

        logger.info(f"上机表写入 {count} 行到 '{self.SHEET_NAME}'")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path


# ------------------------------------------------------------------
# 4. 打包报表 → 数据源表（工序=打包）
# ------------------------------------------------------------------

class PackingFiller:
    """表板打包报表 → 数据源表（工序=打包）。"""

    SHEET_NAME = "数据源表"

    def __init__(self, template_path: Path):
        self.template_path = template_path

    def fill(
            self, results: list[PackingResult], output_path: Path,
            *, species: str = "欧橡", workshop: str = "大锯车间",
    ) -> Path:
        wb = openpyxl.load_workbook(self.template_path)
        ws = _check_sheet(wb, self.SHEET_NAME, self.template_path)
        start = _find_first_empty_row(ws)
        current = start
        count = 0

        for r in results:
            date_obj = _parse_date(r.meta.date)
            for e in r.entries:
                if e.piece_count <= 0:
                    continue
                if date_obj:
                    ws.cell(current, 3, date_obj)                      # 日期
                ws.cell(current, 4, workshop)                          # 车间
                ws.cell(current, 5, "产品")                             # 物料性质
                ws.cell(current, 6, "打包")                             # 工序
                ws.cell(current, 7, species)                           # 木种
                ws.cell(current, 8, "刨切表板")                         # 货物名称
                ws.cell(current, 12, self._map_owner(e.owner))         # 货物所有人
                ws.cell(current, 13, e.package_id)                     # 包号
                ws.cell(current, 14, e.grade)                          # 等级
                ws.cell(current, 15, e.craft)                          # 工艺
                ws.cell(current, 18, e.length_mm)                      # 长度
                ws.cell(current, 19, e.width_mm)                       # 宽度
                ws.cell(current, 20, e.thickness)                      # 厚度
                ws.cell(current, 24, e.piece_count)                    # 片数
                current += 1
                count += 1

        logger.info(f"打包表写入 {count} 行到 '{self.SHEET_NAME}'")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    @staticmethod
    def _map_owner(raw: str) -> str:
        """将手写的所有人名称映射为标准名称。"""
        if "王总" in raw or "王" in raw:
            return "王总公司"
        if "新" in raw or "LKV" in raw.upper():
            return "LKVO新公司"
        return raw
