"""纸质发票记录表 Excel 模板自动填充。

目标模板：纸质发票记录表-20260303.xlsx、Sheet：原始汇总
填充列 (A-L)：
  A: 所属人         — 固定配置值
  B: 公司名         — 固定配置值
  C: 供应商名称     — 从出口商字段提取
  D: 性质           — 进口单据默认“进口”
  E: 名称           — 根据文档类型/货物描述推导
  F: 日期           — 文档日期字段
  G: 发票号         — 发票号或报关单号
  H: 编号           — 自动生成的内部编号
  I: FSC            — “是”或“非”（根据 EUR.1 证书判断）
  J: 供应商/备注    — 公式 =C{row}（自动）
  K: 外币金额       — 从 total_value 提取的数值
  L: 数量           — 从 quantity 提取的数值 (m³)
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import openpyxl
from loguru import logger

from app.schemas import PipelineResult, CustomsRecord

# ---------------------------------------------------------------------------
# 配置：固定列的默认值
# ---------------------------------------------------------------------------

DEFAULT_OWNER = "新A"  # 所属人 — column A
DEFAULT_COMPANY = "AL"  # 公司名 — column B
DEFAULT_NATURE = "进口"  # 性质  — column D
DEFAULT_FSC = "非"  # FSC   — column I


class InvoiceFiller:
    """从管线处理结果填充纸质发票记录表 Excel 模板。"""

    def __init__(
            self,
            template_path: str | Path,
            sheet_name: str = "原始汇总",
            owner: str = DEFAULT_OWNER,
            company: str = DEFAULT_COMPANY,
    ):
        self.template_path = Path(template_path)
        self.sheet_name = sheet_name
        self.owner = owner
        self.company = company

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")

    def fill(
            self,
            results: list[PipelineResult],
            output_path: Path,
            batch_id: Optional[str] = None,
    ) -> Path:
        """使用一个或多个管线结果填充模板。

        参数:
            results: 处理 PDF 文件后的 PipelineResult 列表。
            output_path: 填充后的 Excel 保存路径。
            batch_id: 可选的批次标识前缀，用于内部编号生成。

        返回:
            保存的文件路径。
        """
        wb = openpyxl.load_workbook(self.template_path)
        if self.sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"模板文件 '{self.template_path.name}' 中不存在工作表 '{self.sheet_name}'。"
                f"可用工作表: [{available}]"
            )
        ws = wb[self.sheet_name]

        # 找到第一个空行（表头在第 2 行，数据从第 3 行开始）
        start_row = self._find_first_empty_row(ws)
        logger.info(f"Template has data up to row {start_row - 1}. Appending from row {start_row}.")

        ref_counter = 0
        has_fsc = False  # 跟踪批次中是否发现 EUR.1 证书
        # 第一轮：扫描所有记录，提取 EUR.1 和各报关单的关单号
        # 将关单号与来源文件关联（filename → 关单号）
        customs_refs: dict[str, str] = {}
        for result in results:
            for record in result.records:
                doc_type = self._get_field(record, "document_type").lower()
                if "eur.1" in doc_type or "eur1" in doc_type or "movement certificate" in doc_type:
                    has_fsc = True
                if "customs declaration" in doc_type:
                    ref = self._extract_customs_ref(
                        self._get_field(record, "declaration_number"),
                        self._get_field(record, "remarks"),
                        result.filename,
                    )
                    if ref:
                        customs_refs[result.filename] = ref

        # 第二轮：为相关单据填充行
        # 关单号匹配策略：遇到报关单文件时更新当前关单号，后续文件沿用
        current_customs_ref: str | None = None
        current_row = start_row
        for result in results:
            source_file = result.filename

            # 如果本文件包含报关单，更新当前关单号
            if source_file in customs_refs:
                current_customs_ref = customs_refs[source_file]

            for record in result.records:
                doc_type = self._get_field(record, "document_type").lower()

                # 跳过非发票类单据（续页、仅用于 FSC 判断的证书等）
                if self._should_skip(doc_type):
                    logger.debug(f"Skipping {doc_type} from {source_file} (not billable)")
                    continue

                ref_counter += 1
                ref_num = self._generate_ref(batch_id, record, ref_counter)

                row_data = self._map_record_to_row(record, ref_num, has_fsc)
                # 将当前批次的关单号附到 V 列
                row_data["V"] = current_customs_ref
                self._write_row(ws, current_row, row_data)
                current_row += 1

        records_added = current_row - start_row
        logger.info(f"Added {records_added} row(s) to '{self.sheet_name}'")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        wb.close()
        logger.info(f"Saved filled template to: {output_path}")
        return output_path

    # ------------------------------------------------------------------
    # 映射逻辑
    # ------------------------------------------------------------------

    def _map_record_to_row(
            self, record: CustomsRecord, ref_num: str, has_fsc: bool
    ) -> dict:
        """将 CustomsRecord 映射为列值字典。"""
        doc_type = self._get_field(record, "document_type").lower()
        exporter = self._get_field(record, "exporter")
        importer = self._get_field(record, "importer")
        date_str = self._get_field(record, "date")
        invoice_num = self._get_field(record, "invoice_number")
        decl_num = self._get_field(record, "declaration_number")
        total_value = self._get_field(record, "total_value")
        quantity = self._get_field(record, "quantity")
        goods_desc = self._get_field(record, "goods_description")

        # 智能纠正 importer/exporter 反转
        # 如果 exporter 包含己方公司名（TERRA DRVO），说明 VLM 搞反了
        supplier_raw = self._resolve_supplier(exporter, importer)

        # 确定名称（E 列）
        item_name = self._derive_item_name(doc_type, goods_desc)

        # 确定发票号（G 列）— 优先使用 invoice_number，回退到 declaration_number
        bill_number = invoice_num if invoice_num else decl_num

        # 确定 FSC
        fsc = "是" if has_fsc else DEFAULT_FSC

        # 解析数值
        amount = self._parse_european_number(total_value)
        qty = self._parse_european_number(quantity)

        # 解析日期
        date_val = self._parse_date(date_str)

        # 提取简短供应商名称
        supplier = self._extract_supplier_name(supplier_raw)

        return {
            "A": self.owner,  # 所属人
            "B": self.company,  # 公司名
            "C": supplier,  # 供应商名称
            "D": DEFAULT_NATURE,  # 性质
            "E": item_name,  # 名称
            "F": date_val,  # 日期
            "G": bill_number,  # 发票号
            "H": ref_num,  # 编号
            "I": fsc,  # FSC
            # J is formula =C{row}, handled in _write_row
            "K": amount,  # 外币金额
            "L": qty,  # 数量
        }

    @staticmethod
    def _should_skip(doc_type: str) -> bool:
        """判断该文档类型是否应跳过（非计费项目）。

        报关单 (customs declaration) 也跳过，因为它不单独占行，
        其关单号会附在对应商业发票行的 V 列。
        """
        skip_types = [
            "customs declaration",
            "eur.1", "eur1", "movement certificate",
            "inspection certificate", "phytosanitary",
            "cmr", "transport document",
        ]
        return any(t in doc_type for t in skip_types)

    @staticmethod
    def _resolve_supplier(exporter: str, importer: str) -> str:
        """智能判断供应商：如果 exporter 是己方公司则说明 VLM 反转了字段。

        VLM 有时会把进口发票上的买方（TERRA DRVO）误标为 exporter，
        把卖方误标为 importer。此方法检测并纠正这种情况。
        """
        own_company_keywords = ["terra drvo", "terra drvo doo", "terra drvo d.o.o"]
        exporter_lower = exporter.lower() if exporter else ""

        # 如果 exporter 是己方公司，取 importer 作为供应商
        if any(kw in exporter_lower for kw in own_company_keywords):
            if importer:
                logger.debug(
                    f"importer/exporter 反转纠正: exporter='{exporter[:30]}' → 使用 importer='{importer[:30]}'")
                return importer
        return exporter

    @staticmethod
    def _derive_item_name(doc_type: str, goods_desc: str) -> str:
        """根据文档类型和货物描述推导「名称」列的值。"""
        goods_lower = goods_desc.lower()

        # 运输/运费
        if any(w in doc_type for w in ["transport", "efaktura"]):
            return "运费"
        if any(w in goods_lower for w in ["prevoz", "transport", "运费", "运输"]):
            return "运费"

        # 报关服务发票
        if any(w in goods_lower for w in ["špedicij", "spedicij", "usluga", "报关费", "obracun", "pdv za jci"]):
            return "报关费"

        # 报关单 → 原木
        if "customs" in doc_type or "declaration" in doc_type:
            return "原木"

        # 木材/原木
        if any(w in goods_lower for w in ["trup", "drvo", "原木", "wood", "oak", "hrast", "furnir"]):
            return "原木"

        return "原木"  # 默认

    @staticmethod
    def _extract_supplier_name(exporter: str) -> str:
        """从完整的出口商字符串中提取简短供应商名称。"""
        if not exporter:
            return ""

        name = exporter.strip()

        # 已知供应商映射（长名称 → 短名称）
        supplier_map = {
            "HRVATSKE ŠUME": "HRVATSKE ŠUME",
            "PREMIUM": "PREMIUM",
            "UŠP": "HRVATSKE ŠUME",
            "A.D. GAJ": "AD GAJ",
            "AD GAJ": "AD GAJ",
            "A.D.GAJ": "AD GAJ",
            "KULAS": "KULAS",
            "ALPINA": "ALPINA",
        }
        name_upper = name.upper()
        for key, short in supplier_map.items():
            if key.upper() in name_upper:
                return short

        # 对运输公司，取 "PR " 之前的部分
        if " PR " in name:
            # e.g. "MILENKO JANJATOVIĆ PR AUTOPREVOZ..." → "WOOD TRANS" if found
            if "WOOD TRANS" in name.upper():
                return "WOOD TRANS MOROVIĆ"
            name = name[:name.index(" PR ")]

        # 在地址标识处截断
        for sep in [", Nikol", ", NIKOL", ", Milke", ", MILKE", ", Morović",
                    ", Zagreb", ", ZAGREB", ", Požega", ", POŽEGA",
                    ", USP", ", UŠP", ", UŠĆE", ", Jarački"]:
            if sep in name:
                name = name[:name.index(sep)]
                break

        return name.strip().rstrip(",").strip()

    @staticmethod
    def _parse_european_number(text: str) -> Optional[float]:
        """解析欧洲格式数字（1.234,56）为浮点数。

        欧洲惯例：逗号是小数分隔符，点是千位分隔符。
        但也需处理标准格式（点为小数）和混合情况。
        """
        if not text:
            return None

        raw = text

        # 检测原始文本中是否有体积/重量单位提示（逗号必定是小数分隔符）
        has_metric_unit = bool(re.search(r"m[³3²2]|kg|ton", raw, re.IGNORECASE))

        # 移除币种代码和单位
        cleaned = re.sub(r"[A-Za-z³²%]+", "", raw).strip()
        # 移除空格
        cleaned = cleaned.replace(" ", "")

        if not cleaned:
            return None

        # 欧洲格式：103.700,00 → 103700.00
        if "," in cleaned and "." in cleaned:
            # 确定哪个是小数分隔符
            last_comma = cleaned.rfind(",")
            last_dot = cleaned.rfind(".")
            if last_comma > last_dot:
                # 逗号是小数点：1.234,56
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                # 点是小数点：1,234.56
                cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            # 只有逗号：可能是小数 (3,61 / 23,916) 或千位分隔 (1,000)
            parts = cleaned.split(",")
            if len(parts) == 2:
                after_comma = parts[1]
                # 有体积/重量单位时，逗号一定是小数
                if has_metric_unit:
                    cleaned = cleaned.replace(",", ".")
                # 逗号后 1-2 位：小数（3,61 → 3.61）
                elif len(after_comma) <= 2:
                    cleaned = cleaned.replace(",", ".")
                # 逗号后 3 位且整数部分 < 4 位：更可能是小数（23,916 → 23.916）
                elif len(after_comma) == 3 and len(parts[0]) <= 3:
                    cleaned = cleaned.replace(",", ".")
                else:
                    # 千位分隔（如 103,700）
                    cleaned = cleaned.replace(",", "")
            else:
                cleaned = cleaned.replace(",", "")
        # 其他情况：只有点或无分隔符 — 标准格式

        try:
            return float(cleaned)
        except ValueError:
            return None

    @staticmethod
    def _parse_date(date_str: str) -> Union[datetime, str, None]:
        """解析日期字符串为 datetime。"""
        if not date_str:
            return None

        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue
        return date_str  # 回退：以字符串形式返回

    @staticmethod
    def _generate_ref(batch_id: Optional[str], record: CustomsRecord, counter: int) -> str:
        """生成内部编号（H 列）。"""
        if batch_id:
            return f"{batch_id}-{counter:03d}"
        # 默认：IM + 日期前缀 + 序号
        date_field = ""
        for f in record.fields:
            if f.field_name == "date" and f.value:
                date_field = f.value.replace("-", "")[:6]  # e.g. "202602"
                break
        return f"IM{date_field}{counter:02d}"

    # ------------------------------------------------------------------
    # Excel 读写辅助方法
    # ------------------------------------------------------------------

    @staticmethod
    def _get_field(record: CustomsRecord, field_name: str) -> str:
        """按名称获取记录中的字段值。"""
        for f in record.fields:
            if f.field_name == field_name:
                return f.value
        return ""

    @staticmethod
    def _find_first_empty_row(ws, start: int = 3) -> int:
        """从 `start` 开始找到 A 列的第一个空行。"""
        row = start
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        return row

    @staticmethod
    def _extract_customs_ref(decl_number: str, remarks: str, filename: str = "") -> str | None:
        """从报关单中提取关单号（V 列使用）。

        客户模板中的“关单”是一个简短编号（如 418、467）。
        报关单号格式为 42072C420265659，其中末尾四位数字（5659）是核心编号。
        """
        if not decl_number:
            return None

        # 尝试从标准格式中提取：42072C420265659 → 5659
        cleaned = re.sub(r'[\s\-]', '', decl_number)
        # 匹配 42072C4YYYY#### 格式
        m = re.search(r'42072C4\d{4}(\d{3,5})$', cleaned)
        if m:
            return m.group(1)

        # 尝试从 remarks 中提取 JCI 号
        if remarks:
            jci_match = re.search(r'JCI\d{5,}', remarks)
            if jci_match:
                return jci_match.group()

        # 回退：取末尾 4-5 位数字
        m = re.search(r'(\d{4,5})$', cleaned)
        if m:
            return m.group(1)

        # 回退：从文件名提取（如 Deklaracija_42072_C4_3141_2026 → 3141）
        if filename:
            fn_match = re.search(r'C4[_\s](\d{3,5})', filename)
            if fn_match:
                return fn_match.group(1)

        return decl_number[:20] if decl_number else None

    @staticmethod
    def _write_row(ws, row: int, data: dict):
        """将一行数据写入工作表。"""
        col_map = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6,
                   "G": 7, "H": 8, "I": 9, "K": 11, "L": 12,
                   "V": 22}  # V 列 = 关单号

        # 尽可能复制上一行的格式
        src_row = row - 1 if row > 3 else 3

        for col_letter, col_idx in col_map.items():
            value = data.get(col_letter)
            if value is None:
                continue
            cell = ws.cell(row=row, column=col_idx, value=value)

            # 复制源行的数字格式
            src_cell = ws.cell(row=src_row, column=col_idx)
            if src_cell.number_format:
                cell.number_format = src_cell.number_format

        # J 列（供应商/备注）始终是公式: =C{row}
        ws.cell(row=row, column=10, value=f"=C{row}")
