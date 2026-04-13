"""测试：从管线抽取结果填充纸质发票记录表。

使用一组关联 PDF (c4/5659 批次)：
  - Deklaracija (报关单)
  - Racun (报关公司服务发票)
  - SKEN.36 (扫描合集: 商业发票、EUR.1、检疫证、CMR、电子发票)
"""

import sys
import io
import json
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from app.export.invoice_filler import InvoiceFiller
from app.schemas import PipelineResult

TEMPLATE = Path(__file__).resolve().parent.parent / "文档" / "样例" / "纸质发票记录表-20260303.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# 从 JSON 加载管线处理结果
RESULT_FILES = [
    OUTPUT_DIR / "Deklaracija_42072_C4_5659_2026 [JCI00158583].json",
    OUTPUT_DIR / "Racun.json",
    OUTPUT_DIR / "SKEN.36.json",
]


def load_result(path: Path) -> PipelineResult:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return PipelineResult(**data)


def main():
    print("=== Test: Invoice Template Fill ===\n")

    # 检查文件是否存在
    assert TEMPLATE.exists(), f"Template not found: {TEMPLATE}"
    for rf in RESULT_FILES:
        assert rf.exists(), f"Result not found: {rf}"

    # 加载结果
    results = [load_result(rf) for rf in RESULT_FILES]
    total_records = sum(r.total_documents_detected for r in results)
    print(f"Loaded {len(results)} result files with {total_records} total records")

    # 显示包含的文档
    for r in results:
        print(f"\n  {r.filename}:")
        for rec in r.records:
            doc_type = ""
            for f in rec.fields:
                if f.field_name == "document_type":
                    doc_type = f.value
                    break
            print(f"    Record {rec.record_index} (page {rec.source_page}): {doc_type}")

    # 填充模板
    filler = InvoiceFiller(template_path=TEMPLATE, owner="新A", company="AL")
    output_path = OUTPUT_DIR / "纸质发票记录表-filled.xlsx"
    filler.fill(results, output_path, batch_id="C4-5659")

    # 验证：回读并显示新增行
    import openpyxl
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["原始汇总"]

    print(f"\n{'='*80}")
    print("FILLED ROWS:")
    print(f"{'='*80}")

    # 找到数据开始位置（在已有 18 行之后）
    headers = ["所属人", "公司名", "供应商名称", "性质", "名称", "日期",
               "发票号", "编号", "FSC", "供应商/备注", "外币金额", "数量"]

    for row_idx in range(3, ws.max_row + 1):
        val_a = ws.cell(row=row_idx, column=1).value
        if val_a is None:
            break

    # 显示新增行（从第 21 行开始，因为模板有 18 数据行 + 2 表头行）
    new_start = 21  # row 3 + 18 existing = row 21
    for row_idx in range(new_start, ws.max_row + 1):
        val_a = ws.cell(row=row_idx, column=1).value
        if val_a is None:
            break
        print(f"\n  Row {row_idx}:")
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                print(f"    {header}: {val}")

    wb.close()
    print(f"\n✓ Output saved to: {output_path}")


if __name__ == "__main__":
    main()
