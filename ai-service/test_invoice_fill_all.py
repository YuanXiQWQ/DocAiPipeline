"""测试：全批次纸质发票记录表填充。

覆盖三个进口批次 (3141, 4143, 5659)，每个批次包含：
  - Deklaracija (报关单) → 跳过，仅提取关单号到 V 列
  - SKEN (扫描合集) → 商业发票 + 运费发票
  - CR (税款通知) → 暂不填入
"""

import sys
import io
import json
from pathlib import Path
from typing import TypedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from app.export.invoice_filler import InvoiceFiller
from app.schemas import PipelineResult

TEMPLATE = Path(__file__).resolve().parent.parent / "文档" / "样例" / "纸质发票记录表-20260303.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"


def load_result(path: Path) -> PipelineResult:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return PipelineResult(**data)


class _Batch(TypedDict):
    id: str
    files: list[Path]


# 按批次组织文件（每批次：报关单 + SKEN）
BATCHES: list[_Batch] = [
    {
        "id": "C4-3141",
        "files": [
            OUTPUT_DIR / "Deklaracija_42072_C4_3141_2026 [JCI00158457].json",
            OUTPUT_DIR / "SKEN.52.json",
        ],
    },
    {
        "id": "C4-4143",
        "files": [
            OUTPUT_DIR / "Deklaracija_42072_C4_4143_2026 [JCI00158495].json",
            OUTPUT_DIR / "SKEN.78.json",
        ],
    },
    {
        "id": "C4-5659",
        "files": [
            OUTPUT_DIR / "Deklaracija_42072_C4_5659_2026 [JCI00158583].json",
            OUTPUT_DIR / "Racun.json",
            OUTPUT_DIR / "SKEN.36.json",
        ],
    },
]


def main():
    print("=== 全批次纸质发票记录表填充测试 ===\n")

    assert TEMPLATE.exists(), f"模板未找到: {TEMPLATE}"

    # 收集所有结果
    all_results = []
    for batch in BATCHES:
        print(f"\n--- 批次 {batch['id']} ---")
        for fp in batch["files"]:
            if not fp.exists():
                print(f"  ⚠ 跳过: {fp.name} (文件不存在)")
                continue
            result = load_result(fp)
            all_results.append(result)
            doc_types = []
            for rec in result.records:
                for f in rec.fields:
                    if f.field_name == "document_type":
                        doc_types.append(f.value)
                        break
            print(f"  ✓ {result.filename}: {len(result.records)} 记录 ({', '.join(doc_types)})")

    print(f"\n共加载 {len(all_results)} 个结果文件")

    # 填充模板
    filler = InvoiceFiller(template_path=TEMPLATE, owner="新A", company="AL")
    output_path = OUTPUT_DIR / "纸质发票记录表-全批次.xlsx"
    filler.fill(all_results, output_path)

    # 验证结果
    import openpyxl
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["原始汇总"]

    print(f"\n{'=' * 100}")
    print("填充结果（新增行）:")
    print(f"{'=' * 100}")
    print(f"{'行':>4s}  {'供应商':15s}  {'名称':6s}  {'发票号':20s}  {'金额':>12s}  {'数量':>8s}  {'关单':>6s}  {'FSC':>4s}")
    print(f"{'-' * 4}  {'-' * 15}  {'-' * 6}  {'-' * 20}  {'-' * 12}  {'-' * 8}  {'-' * 6}  {'-' * 4}")

    for row_idx in range(21, ws.max_row + 1):
        val_a = ws.cell(row=row_idx, column=1).value
        if val_a is None:
            break
        supplier = ws.cell(row=row_idx, column=3).value or ""
        name = ws.cell(row=row_idx, column=5).value or ""
        inv = ws.cell(row=row_idx, column=7).value or ""
        amount = ws.cell(row=row_idx, column=11).value or ""
        qty = ws.cell(row=row_idx, column=12).value or ""
        gd = ws.cell(row=row_idx, column=22).value or ""
        fsc = ws.cell(row=row_idx, column=9).value or ""
        print(f"{row_idx:>4d}  {str(supplier):15s}  {str(name):6s}  {str(inv):20s}  {str(amount):>12s}  {str(qty):>8s}  {str(gd):>6s}  {str(fsc):>4s}")

    wb.close()
    print(f"\n✓ 输出保存至: {output_path}")


if __name__ == "__main__":
    main()
